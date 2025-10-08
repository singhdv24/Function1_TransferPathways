import gradio as gr
import pandas as pd
import os, re, tempfile, unicodedata
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------- cleaning helpers ----------
def clean_text(s: str) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    # normalize unicode, replace non-breaking spaces and weird dashes
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace("\u00A0", " ")  # NBSP -> space
    s = s.replace("\u2014", "-").replace("\u2013", "-")  # em/en dash -> hyphen
    # collapse whitespace
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def normalize_code(code):
    c = clean_text(code).upper()
    return c if c else None

def find_col(df, candidates, label):
    for c in df.columns:
        low = c.lower()
        for want in candidates:
            if want in low:
                return c
    raise ValueError(f"{label} column not found. Looked for any of {candidates} in {list(df.columns)}")

def safe_token(s):
    s = clean_text(s)
    s = re.sub(r"[^\w]+", " ", s).strip()
    parts = s.split()
    bad = {"university","college","community","of","the","district","cc"}
    parts = [p for p in parts if p.lower() not in bad]
    if not parts:
        parts = s.split()
    return "".join(parts[:2])  # keep it short and filesystem-safe

def infer_inst_plan_from_filename(path):
    base = os.path.splitext(os.path.basename(path))[0]
    parts = base.split("_", 2)
    if len(parts) >= 3 and parts[0].upper() in ("AS","BS"):
        inst = safe_token(parts[1])
        plan = re.sub(r"\s+", "", parts[2])
        return inst, plan
    return safe_token(base), "Plan"

# ---------- loaders ----------
def load_as(df):
    code_col   = find_col(df, ["course_name","course code","course_code","name","code"], "AS course-code")
    credit_col = find_col(df, ["credit_hours","credits","credit hours"], "AS credits")
    term_col   = next((c for c in df.columns if "term" in c.lower()), None)

    out = pd.DataFrame({
        "AS_Code": df[code_col].apply(normalize_code),
        "AS_Credits": pd.to_numeric(df[clean_text(credit_col)] if credit_col in ["", None] else df[credit_col], errors="coerce")
    })
    out["AS_Term"] = df[term_col] if term_col else range(1, len(out)+1)
    out = out.dropna(subset=["AS_Code","AS_Credits"])
    return out[["AS_Term","AS_Code","AS_Credits"]]

def load_bs(df):
    code_col   = find_col(df, ["name","course_name","course code","course_code","code"], "BS course-code")
    credit_col = find_col(df, ["credits","credit","credit hours","credit_hours"], "BS credits")
    term_col   = next((c for c in df.columns if "term" in c.lower()), None)

    out = pd.DataFrame({
        "BS_Code": df[code_col].apply(normalize_code),
        "BS_Credits": pd.to_numeric(df[credit_col], errors="coerce")
    })
    out["BS_Term"] = df[term_col] if term_col else range(1, len(out)+1)
    out = out.dropna(subset=["BS_Code"])
    return out[["BS_Term","BS_Code","BS_Credits"]]

def load_equiv(df):
    as_col = find_col(df, ["course_code","as_code"], "Equivalency AS code")
    eq_col = "Equivalent Course Code"
    if eq_col not in df.columns:
        eq_col = find_col(df, ["equivalent course code","equivalent"], "Equivalency BS code")
    out = pd.DataFrame({
        "AS_Code": df[as_col].apply(normalize_code),
        "BS_Codes": df[eq_col].astype(str).str.split(";").apply(lambda xs: [normalize_code(x) for x in xs if x])
    }).dropna(subset=["AS_Code"])
    return out

# ---------- write formatted xlsx ----------
def write_formatted_xlsx(df: pd.DataFrame, as_inst, as_plan, bs_inst, bs_plan) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "combined_plan"

    # write dataframe to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append([clean_text(v) for v in r])

    # styles
    wrap = Alignment(wrap_text=True, vertical="top")
    header_font = Font(bold=True)
    colA_font = Font(bold=True)  # make column A (Term) bold

    # apply wrap to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    # header row bold
    for cell in ws[1]:
        cell.font = header_font

    # column A bold (Term)
    for r in range(1, ws.max_row + 1):
        ws[f"A{r}"].font = colA_font

    # freeze header
    ws.freeze_panes = "A2"

    # autosize columns (simple heuristic)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # save to temp
    tmpdir = tempfile.gettempdir()
    fname = f"combined_study_plan_AS_{as_inst}_{as_plan}__BS_{bs_inst}_{bs_plan}.xlsx"
    out_path = os.path.join(tmpdir, fname)
    wb.save(out_path)
    return out_path

# ---------- core combine ----------
def combine(as_xlsx, bs_xlsx, equiv_xlsx):
    as_path, bs_path, eq_path = as_xlsx.name, bs_xlsx.name, equiv_xlsx.name
    as_inst, as_plan = infer_inst_plan_from_filename(as_path)
    bs_inst, bs_plan = infer_inst_plan_from_filename(bs_path)

    as_df = load_as(pd.read_excel(as_path))
    bs_df = load_bs(pd.read_excel(bs_path))
    eq_df = load_equiv(pd.read_excel(eq_path, sheet_name=0))

    eq_map = {r.AS_Code: set(r.BS_Codes) for _, r in eq_df.iterrows()}
    bs_codes = set(bs_df["BS_Code"])

    combined_rows = []
    matched_bs_codes = set()

    for _, row in as_df.iterrows():
        as_code, as_cr, term = row["AS_Code"], float(row["AS_Credits"]), row["AS_Term"]
        bs_equivs = eq_map.get(as_code, set())
        hits = [b for b in bs_equivs if b in bs_codes]
        if hits:
            bs_hit = hits[0]
            matched_bs_codes.add(bs_hit)
            combined_rows.append([term, "AS", "✅", as_code, as_cr, bs_hit, "Transferred"])
        else:
            combined_rows.append([term, "AS", "❌", as_code, as_cr, "", "Not transferred"])

    remaining_bs = bs_df[~bs_df["BS_Code"].isin(matched_bs_codes)]
    for _, r in remaining_bs.iterrows():
        combined_rows.append([r["BS_Term"], "BS", "", "", "", r["BS_Code"], "To complete at BS"])

    out = pd.DataFrame(
        combined_rows,
        columns=["Term","Source","Match","AS Course","AS Credits","BS Course","Status"]
    ).sort_values(by=["Term","Source"], kind="stable")

    # produce formatted xlsx
    xlsx_path = write_formatted_xlsx(out, as_inst, as_plan, bs_inst, bs_plan)
    return out, xlsx_path

# ---------- UI ----------
demo = gr.Interface(
    fn=combine,
    inputs=[
        gr.File(label="AS Plan (.xlsx)"),
        gr.File(label="BS Plan (.xlsx)"),
        gr.File(label="Course Equivalencies (.xlsx)")
    ],
    outputs=[
        gr.Dataframe(label="Combined Plan of Study"),
        gr.File(label="Download Excel (.xlsx)")
    ],
    title="Combine Plans of Study",
    description="Uploads one AS plan, one BS plan, and equivalencies; outputs a formatted, combined study plan."
)

if __name__ == "__main__":
    demo.launch()
